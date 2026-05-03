"""
엑셀 그리드 방식 — 작은 셀(2pt×2pt) 격자에 PDF 텍스트/표 위치 정확 매핑.
A4 페이지 크기 fit-to-page 로 PDF 와 시각 동일.
"""
import sys, time
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from playwright.sync_api import sync_playwright
import win32com.client as wc
import fitz
from PIL import Image, ImageChops
try: sys.stdout.reconfigure(encoding="utf-8")
except: pass

PDF = Path(r"C:\Users\신현식\Desktop\1.판교글로벌비즈센터 산업시설(B-301~303호) 처분 수의계약 공고문.pdf")
OUT = Path(__file__).parent / "out_excel_grid"; OUT.mkdir(exist_ok=True)

URL = "https://hyshin6664.github.io/hwpx-editor/?cb=" + str(int(time.time()*1000))

def collect_pdf_data():
    """우리 앱으로 PDF 마운트 후 textBoxes/lineRects/fillRects 추출"""
    with sync_playwright() as pw:
        b = pw.chromium.launch(headless=True)
        page = b.new_context(viewport={"width": 1280, "height": 900}).new_page()
        page.on("dialog", lambda d: d.accept())
        page.goto(URL, wait_until="domcontentloaded", timeout=60000)
        page.wait_for_function("() => window.__editorReady === true", timeout=60000)
        page.set_input_files("#picker", str(PDF))
        page.wait_for_function("() => window.__currentMode === 'pdf'", timeout=120000)
        page.wait_for_timeout(2500)
        data = page.evaluate("""() => {
          return window.__pdfState.pages.map(pg => ({
            pageW: pg.pdfPageWidth,
            pageH: pg.pdfPageHeight,
            scale: pg.scale,
            textBoxes: pg.textBoxes || [],
            lineRects: pg.lineRects || [],
            fillRects: pg.fillRects || [],
          }));
        }""")
        b.close()
    return data

def build_excel(pages):
    """각 페이지 → worksheet, 작은 셀 격자 + 텍스트 배치 + 표 테두리"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    # 그리드 해상도 — 매우 미세 (1pt 단위)
    GRID_W = 595  # A4 595pt → 1pt per cell
    for pi, pg in enumerate(pages):
        # canvas 좌표계 — pg.scale 곱한 px. PDF 좌표(pt) = canvas / scale
        scale = pg['scale']
        canvas_w = pg['pageW'] * scale
        canvas_h = pg['pageH'] * scale
        cell_canvas_w = canvas_w / GRID_W
        # 행 높이 — 1pt = 점 단위
        ROW_H_PT = 1
        rows_per_page = int(pg['pageH'] / ROW_H_PT)  # A4 ≈ 842 행

        ws = wb.create_sheet(f"Page {pi+1}")
        # 컬럼 너비 — Excel 컬럼 width 단위는 character. 1 char ≈ 7px ≈ 5.25pt
        # cell_canvas_w / scale = pt 단위. 1pt = 1/5.25 char
        col_width_chars = (cell_canvas_w / scale) / 5.25
        for c in range(1, GRID_W + 1):
            ws.column_dimensions[get_column_letter(c)].width = col_width_chars
        for r in range(1, rows_per_page + 1):
            ws.row_dimensions[r].height = ROW_H_PT  # pt 단위

        # 텍스트 배치 — 미세 그리드 충돌 방지
        small_font = Font(name='맑은 고딕', size=7)
        center = Alignment(horizontal='left', vertical='center', wrap_text=False)
        # 사용된 (row, col) 추적
        used = set()
        # textBoxes 정렬 (위→아래, 왼쪽→오른쪽)
        sorted_tbs = sorted(pg['textBoxes'], key=lambda t: (t['y'], t['x']))
        for tb in sorted_tbs:
            col = int(tb['x'] / cell_canvas_w) + 1
            row = int(tb['y'] / (ROW_H_PT * scale)) + 1
            if col < 1 or col > GRID_W or row < 1 or row > rows_per_page: continue
            col_span = max(1, int(tb['w'] / cell_canvas_w))
            end_col = min(GRID_W, col + col_span - 1)
            # 충돌 검사 — 이미 사용된 셀은 다음 비어있는 위치로 이동
            while col <= GRID_W and (row, col) in used:
                col += 1
            if col > GRID_W: continue
            try:
                cell = ws.cell(row=row, column=col, value=tb['str'])
                cell.font = small_font
                cell.alignment = center
                if end_col > col:
                    end_col_safe = end_col
                    # merge 영역에 이미 사용된 셀이 있으면 영역 줄임
                    while end_col_safe > col and (row, end_col_safe) in used:
                        end_col_safe -= 1
                    if end_col_safe > col:
                        try:
                            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col_safe)
                            for c in range(col, end_col_safe + 1):
                                used.add((row, c))
                        except Exception:
                            used.add((row, col))
                    else:
                        used.add((row, col))
                else:
                    used.add((row, col))
            except Exception:
                pass

        # 표 테두리 — lineRects 의 가는 사각형(=세로/가로 선)
        thin = Side(style='thin', color='000000')
        def safe_border(r, c, **kw):
            try:
                cell = ws.cell(row=r, column=c)
                if cell.__class__.__name__ == 'MergedCell': return
                old = cell.border
                cell.border = Border(
                    left=kw.get('left', old.left),
                    top=kw.get('top', old.top),
                    right=kw.get('right', old.right),
                    bottom=kw.get('bottom', old.bottom),
                )
            except Exception: pass
        for lr in pg['lineRects']:
            x = lr['x']; y = lr['y']; w = lr['w']; h = lr['h']
            if w < 4 and h > 20:
                col = int(x / cell_canvas_w) + 1
                row1 = max(1, int(y / (ROW_H_PT * scale)) + 1)
                row2 = min(rows_per_page, int((y + h) / (ROW_H_PT * scale)) + 1)
                for r in range(row1, row2 + 1):
                    safe_border(r, col, left=thin)
            elif h < 4 and w > 20:
                row = int(y / (ROW_H_PT * scale)) + 1
                col1 = max(1, int(x / cell_canvas_w) + 1)
                col2 = min(GRID_W, int((x + w) / cell_canvas_w) + 1)
                for c in range(col1, col2 + 1):
                    safe_border(row, c, top=thin)

        # 셀 배경색 — fillRects (회색 헤더 등)
        for fr in pg['fillRects']:
            r0, g0, b0 = fr['color']
            if r0 >= 248 and g0 >= 248 and b0 >= 248: continue  # 흰색 skip
            hex_color = '%02X%02X%02X' % (r0, g0, b0)
            col1 = max(1, int(fr['x'] / cell_canvas_w) + 1)
            col2 = min(GRID_W, int((fr['x'] + fr['w']) / cell_canvas_w) + 1)
            row1 = max(1, int(fr['y'] / (ROW_H_PT * scale)) + 1)
            row2 = min(rows_per_page, int((fr['y'] + fr['h']) / (ROW_H_PT * scale)) + 1)
            for r in range(row1, row2 + 1):
                for c in range(col1, col2 + 1):
                    try:
                        cell = ws.cell(row=r, column=c)
                        if cell.__class__.__name__ != 'MergedCell':
                            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
                    except Exception: pass

        # 페이지 설정 — A4, 인쇄 시 1페이지에 맞춤
        ws.page_setup.paperSize = ws.PAPERSIZE_A4  # 9
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.3, bottom=0.3)

    out = OUT / "grid_approach.xlsx"
    wb.save(out)
    return out

def excel_to_pdf(xlsx_path):
    """Excel COM 으로 xlsx → pdf"""
    pdf_out = xlsx_path.with_suffix('.pdf')
    excel = wc.Dispatch("Excel.Application")
    excel.Visible = False
    try:
        wb = excel.Workbooks.Open(str(xlsx_path.absolute()))
        wb.ExportAsFixedFormat(0, str(pdf_out.absolute()))  # xlTypePDF=0
        wb.Close(SaveChanges=False)
    finally:
        excel.Quit()
    return pdf_out

def compare_pdfs(orig, new):
    def render(p):
        d = fitz.open(p); imgs = []
        for pg in d:
            pix = pg.get_pixmap(matrix=fitz.Matrix(120/72, 120/72))
            imgs.append(Image.frombytes("RGB", [pix.width, pix.height], pix.samples))
        d.close(); return imgs
    a = render(orig); b = render(new)
    diffs = []
    for i in range(min(len(a), len(b))):
        ia, ib = a[i], b[i]
        if ia.size != ib.size: ib = ib.resize(ia.size)
        diff_px = sum(1 for px in ImageChops.difference(ia, ib).getdata() if any(v > 30 for v in px))
        diffs.append(diff_px / (ia.size[0]*ia.size[1]) * 100)
    return {'avg': sum(diffs)/len(diffs) if diffs else 100, 'orig': len(a), 'new': len(b), 'per_page': diffs}

def main():
    print("📊 1) PDF 데이터 수집 (textBoxes/lineRects/fillRects)")
    pages = collect_pdf_data()
    print(f"   {len(pages)} 페이지")
    for i, p in enumerate(pages):
        print(f"   page {i+1}: textBoxes={len(p['textBoxes'])}, lineRects={len(p['lineRects'])}, fillRects={len(p['fillRects'])}")

    print("\n📊 2) 엑셀 그리드 방식 빌드")
    xlsx = build_excel(pages)
    print(f"   → {xlsx.name} ({xlsx.stat().st_size/1024:.1f} KB)")

    print("\n📊 3) Excel → PDF (Excel COM)")
    pdf2 = excel_to_pdf(xlsx)
    print(f"   → {pdf2.name} ({pdf2.stat().st_size/1024:.1f} KB)")

    print("\n📊 4) 원본 PDF 와 비교")
    r = compare_pdfs(PDF, pdf2)
    print(f"   원본 {r['orig']}p, 변환 {r['new']}p")
    print(f"   평균 차이: {r['avg']:.2f}%")
    for i, d in enumerate(r['per_page']):
        print(f"     page {i+1}: {d:.2f}%")

if __name__ == "__main__":
    main()
